#!/usr/bin/env python3
"""
Version Sync Script for DocLang Standard

This script reads the current doclang version (same resolution as `doclang --version`)
and syncs it to:
- iso-standard.md
- doclang/doclang.xsd
- reference/input/reference.xlsx

The package version is derived from git tags via setuptools-scm (see pyproject.toml).
Tags should follow vMAJOR.MINOR.PATCH (e.g. v0.3.0). Commits after a tag get a
+g<sha> local suffix (e.g. 0.3.0+g93c2a53).

Usage:
    python utils/sync_version.py              # version from git / installed package
    python utils/sync_version.py 0.4.0        # explicit target release version
    python utils/sync_version.py v0.4.0       # v-prefix accepted
"""

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from doclang.version import (
    normalize_version,
    release_version_triple,
    resolve_version,
    validate_version,
)


def resolve_sync_version(version_arg: str | None) -> str:
    """Explicit CLI version, or release triple from the same source as doclang --version."""
    if version_arg is not None:
        try:
            return normalize_version(version_arg)
        except ValueError as exc:
            print(f"Error: {exc}")
            sys.exit(1)

    current = resolve_version()
    if current == "unknown" or not validate_version(current):
        print(f"Error: Could not resolve version (got '{current}')")
        print("Run from a git checkout, `uv sync`, or pass an explicit version (e.g. 0.4.0)")
        sys.exit(1)

    return release_version_triple(current)


def sync_version_in_iso_standard(file_path: Path, version: str) -> None:
    """Update version references in iso-standard.md"""
    content = file_path.read_text(encoding='utf-8')
    
    # Extract MAJOR.MINOR for iso-standard.md
    major_minor = '.'.join(version.split('.')[:2])
    
    # Pattern: "The version of the present specification is **0.2**."
    content = re.sub(
        r'(The version of the present specification is \*\*)\d+\.\d+(\*\*\.)',
        rf'\g<1>{major_minor}\g<2>',
        content
    )
    
    # Note: Attribute definitions are now managed via Excel automation
    # and exported to iso-standard.md, so we don't update them here
    
    file_path.write_text(content, encoding='utf-8')


def sync_version_in_xsd(file_path: Path, version: str) -> None:
    """Update version in doclang.xsd"""
    content = file_path.read_text(encoding='utf-8')
    
    # Extract MAJOR.MINOR for most places
    major_minor = '.'.join(version.split('.')[:2])
    
    # Pattern 1: schema version attribute (line 8) in <xs:schema>
    # version="0.2.0" - full MAJOR.MINOR.PATCH
    # Match only the xs:schema element's version attribute
    content = re.sub(
        r'(<xs:schema[^>]*\s+version=")[\d.]+(")',
        rf'\g<1>{version}\g<2>',
        content,
        flags=re.DOTALL
    )
    
    # Pattern 2: doclang element default version (line 621)
    # <xs:attribute name="version" use="optional" default="0.2">
    # Uses MAJOR.MINOR only
    content = re.sub(
        r'(<xs:attribute name="version" use="optional" default=")[\d.]+(")',
        rf'\g<1>{major_minor}\g<2>',
        content
    )
    
    # Pattern 3: enumeration values (line 627)
    # ALWAYS remove all previous v0.x enumerations and add only the new version
   
    # Remove all enumeration lines (including any trailing whitespace before closing tag)
    content = re.sub(r'\s*<xs:enumeration value="[\d.]+"/>\s*', '', content)
    
    # Add only the new version with proper formatting (MAJOR.MINOR only)
    # Find the comment and the closing </xs:restriction> tag
    insert_pattern = r'(<!-- enumeration set up for v1 and above.*?-->)\s*(</xs:restriction>)'
    content = re.sub(
        insert_pattern,
        rf'\g<1>\n            <xs:enumeration value="{major_minor}"/>\n          \g<2>',
        content,
        flags=re.DOTALL
    )  
    file_path.write_text(content, encoding='utf-8')


def sync_version_in_excel(file_path: Path, version: str) -> None:
    """Update version in reference Excel file"""
    if not file_path.exists():
        print(f"⚠ Skipping {file_path} (file not found)")
        return
    
    # Extract MAJOR.MINOR for Excel
    major_minor = '.'.join(version.split('.')[:2])
    
    try:
        wb = load_workbook(file_path)
        
        # Update the 'attributes' sheet
        if 'attributes' in wb.sheetnames:
            ws = wb['attributes']
            
            # Find the row with `<doclang>` version attribute
            updated = False
            for row in ws.iter_rows(min_row=2):  # Skip header row
                element_cell = row[0]
                attribute_cell = row[1]
                required_cell = row[2]
                values_cell = row[3]
                
                if (element_cell.value == '`<doclang>`' and 
                    attribute_cell.value == '`version`'):
                    
                    # Update "Optional; default: "0.2"" to new version
                    if required_cell.value:
                        required_cell.value = re.sub(
                            r'(Optional; default: ")\d+\.\d+(")',
                            rf'\g<1>{major_minor}\g<2>',
                            str(required_cell.value)
                        )
                    
                    # Update {"0.2"} to new version
                    if values_cell.value:
                        values_cell.value = re.sub(
                            r'(\{")\d+\.\d+("\})',
                            rf'\g<1>{major_minor}\g<2>',
                            str(values_cell.value)
                        )
                    
                    updated = True
                    break
            
            if updated:
                wb.save(file_path)
            else:
                print(f"⚠ Could not find version attribute row in {file_path}")
        else:
            print(f"⚠ 'attributes' sheet not found in {file_path}")
            
    except Exception as e:
        print(f"⚠ Error updating {file_path}: {e}")


def sync_version(version_arg: str | None = None, project_root: Path | None = None) -> str:
    """Sync version across iso-standard.md, doclang.xsd, and reference.xlsx."""
    if project_root is None:
        project_root = Path(__file__).resolve().parent.parent

    iso_standard_path = project_root / "iso-standard.md"
    xsd_path = project_root / "doclang" / "doclang.xsd"
    excel_path = project_root / "reference" / "input" / "reference.xlsx"

    version = resolve_sync_version(version_arg)

    if not iso_standard_path.exists():
        raise FileNotFoundError(f"{iso_standard_path} not found")
    if not xsd_path.exists():
        raise FileNotFoundError(f"{xsd_path} not found")

    sync_version_in_iso_standard(iso_standard_path, version)
    sync_version_in_xsd(xsd_path, version)
    sync_version_in_excel(excel_path, version)

    print(f"Version synced to: {version}")
    return version


def main():
    parser = argparse.ArgumentParser(
        description="Sync DocLang version across iso-standard.md, doclang.xsd, and reference.xlsx",
    )
    parser.add_argument(
        "version",
        nargs="?",
        help="Target release version (e.g. 0.4.0). Defaults to the current doclang version.",
    )
    args = parser.parse_args()

    try:
        sync_version(args.version)
    except FileNotFoundError as exc:
        print(f"Error: {exc}")
        sys.exit(1)


if __name__ == '__main__':
    main()
